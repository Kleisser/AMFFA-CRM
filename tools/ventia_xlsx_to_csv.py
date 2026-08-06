#!/usr/bin/env python3
"""Convierte el reporte de contactos de Ventia (.xlsx) a CSV UTF-8 para importar al CRM.

Uso: python tools/ventia_xlsx_to_csv.py Contactos-reporte-ventia.xlsx backend/storage/app/ventia_contactos.csv
"""
import csv
import re
import sys
from datetime import datetime

import openpyxl

DATE_RE = re.compile(r"^(\d{2})/(\d{2})/(\d{4})(?: (\d{2}):(\d{2})(?::(\d{2}))?)?$")


def normalize(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value)
    m = DATE_RE.match(text)
    if m:
        day, month, year = m.group(1), m.group(2), m.group(3)
        time_part = ""
        if m.group(4):
            time_part = f" {m.group(4)}:{m.group(5)}:{m.group(6) or '00'}"
        return f"{year}-{month}-{day}{time_part}"
    return text


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    src, dst = sys.argv[1], sys.argv[2]

    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]

    with open(dst, "w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            writer.writerow([normalize(v) for v in row])

    print(f"OK: {dst} generado")


if __name__ == "__main__":
    main()
